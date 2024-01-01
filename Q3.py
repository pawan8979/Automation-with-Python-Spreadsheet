# openpyxl package
import openpyxl;

inv_file = openpyxl.load_workbook("inventory.xlsx"); 
product_list = inv_file["Sheet1"];

# list products with inventory less than 10

products_under_10_inv={};

print(product_list.max_row)

for product_row in range(2, product_list.max_row + 1):
    inventory = product_list.cell(product_row, 2).value;
    product_num = product_list.cell(product_row, 1).value;
    
    if inventory < 10:
        products_under_10_inv[int(product_num)] = int(inventory);
    
print(products_under_10_inv);



