# openpyxl package
import openpyxl;

inv_file = openpyxl.load_workbook("inventory.xlsx"); 
product_list = inv_file["Sheet1"];

# calculate and write inventory value for each product into spreadsheet

print(product_list.max_row)

for product_row in range(2, product_list.max_row + 1):
    inventory = product_list.cell(product_row, 2).value;
    price = product_list.cell(product_row, 3).value;
    inventory_price = product_list.cell(product_row, 5);
    
    inventory_price.value = inventory * price;  

inv_file.save("inventory_with_total_value.xlsx");



