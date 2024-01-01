# openpyxl package
import openpyxl;

inv_file = openpyxl.load_workbook("inventory.xlsx"); 
product_list = inv_file["Sheet1"];

# list each company with respective total inventory value

total_value_per_supplier={};

print(product_list.max_row)

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value;
    inventory = product_list.cell(product_row, 2).value;
    price = product_list.cell(product_row, 3).value
    
    if supplier_name in total_value_per_supplier:
        total_value_per_supplier[supplier_name] += inventory*price;
    else:
        total_value_per_supplier[supplier_name] = inventory * price;

print(total_value_per_supplier);



