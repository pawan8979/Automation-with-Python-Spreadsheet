# openpyxl package
import openpyxl;

inv_file = openpyxl.load_workbook("inventory.xlsx"); 
product_list = inv_file["Sheet1"];

# list each company with respective product count

products_per_supplier= {};

print(product_list.max_row)

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value;
    if supplier_name in products_per_supplier:
        print("Increasing count of existing Supplier");
        products_per_supplier[supplier_name] += 1 ;
    else:
        print("Adding a new Supplier");
        products_per_supplier[supplier_name] = 1 ;

print(products_per_supplier);